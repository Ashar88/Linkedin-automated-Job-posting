import gspread
import pandas as pd
import openpyxl

def Excel_Data():
    gc = gspread.oauth()
    sh = gc.open("TRP Jobs Roster Team 3")
    spreadsheet_data = gc.export(sh.id,gspread.utils.ExportFormat.EXCEL)
   
    FILE_NAME = "D:/WORK/All_Python_Work/Python Work/PANDAS folder/Qureos/Playwright/FILES/Roster_File.xlsx"
    SHEET = ""
    FULLNAME = "Ashar Saleem"

    # Write the binary data to a file
    with open(FILE_NAME, 'wb') as file:
        file.write(spreadsheet_data)
        
    wb = openpyxl.load_workbook(FILE_NAME)
    print(wb)
    sheets = wb.sheetnames
    latestWeek = sheets[sheets.index('Dashboards') - 1]
    SHEET = latestWeek

    df = pd.read_excel(FILE_NAME, sheet_name = SHEET)
    
    ws = wb[SHEET]
    columns = list(df.columns)

    for index, row in enumerate(ws.iter_rows()):
        for col, cell in enumerate(row):
            if cell.hyperlink:
                df.loc[index-1, f'col-{  columns[col] }'] = cell.hyperlink.target
                
    mask = df['Recruiter'].apply(lambda x: str(x).lower() == FULLNAME.lower() or str(x).lower() == FULLNAME.lower().split()[0])
    df = df[mask]
    df.to_csv('D:/WORK/All_Python_Work/Python Work/PANDAS folder/Qureos/Playwright/FILES/My_Jobs_Details.csv')

Excel_Data()